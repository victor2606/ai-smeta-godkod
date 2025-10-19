"""
Excel Loader for Construction Rates Data
Reads and validates Excel files with construction rate schedules.
Optimized for large files (100MB+) using chunked reading.
"""

import pandas as pd
import logging
import tempfile
import csv
from pathlib import Path
from typing import Dict, Any, Optional
from tqdm import tqdm
from openpyxl import load_workbook


logger = logging.getLogger(__name__)


class ExcelLoader:
    """
    Loads and validates Excel files containing construction rate data.

    Attributes:
        file_path: Path to Excel file
        df: Loaded DataFrame
    """

    # Required columns that must exist in Excel
    REQUIRED_COLUMNS = [
        'Расценка | Код',
        'Расценка | Исходное наименование',
        'Расценка | Ед. изм.',
        'Тип строки',
        'Ресурс | Код',
        'Ресурс | Стоимость (руб.)',
        'Прайс | АбстРесурс | Сметная цена текущая_median'
    ]

    # Columns that cannot have NaN values
    CRITICAL_COLUMNS = ['Расценка | Код', 'Тип строки']

    def __init__(self, file_path: str, chunk_size: int = 10000):
        """
        Initialize with file path and optional chunk size.

        Args:
            file_path: Path to Excel file
            chunk_size: Number of rows to process per chunk (default: 10000)
        """
        self.file_path = Path(file_path)
        self.chunk_size = chunk_size
        self.df: pd.DataFrame = None
        self._temp_csv_path: Optional[Path] = None

    def _get_total_rows(self) -> int:
        """
        Fast row count without loading entire file.
        Uses openpyxl to read only dimensions.

        Returns:
            Total number of rows (excluding header)
        """
        try:
            wb = load_workbook(self.file_path, read_only=True, data_only=True)
            ws = wb.active
            max_row = ws.max_row
            wb.close()
            return max(0, max_row - 1)  # Subtract header row
        except Exception as e:
            logger.warning(f"Could not determine row count: {e}. Using fallback.")
            return 0  # Fallback to indeterminate progress

    def _convert_xlsx_to_csv(self, temp_csv: Path, total_rows: int) -> None:
        """
        Convert XLSX to CSV with progress tracking.
        This is faster than pd.read_excel for large files.

        Args:
            temp_csv: Path to temporary CSV file
            total_rows: Total rows for progress bar
        """
        logger.info("Converting XLSX to CSV for faster loading...")

        wb = load_workbook(self.file_path, read_only=True, data_only=True)
        ws = wb.active

        with open(temp_csv, 'w', newline='', encoding='utf-8') as csvfile:
            csv_writer = csv.writer(csvfile)

            # Progress bar for conversion
            desc = "Converting XLSX"
            pbar = tqdm(ws.iter_rows(values_only=True),
                       desc=desc,
                       total=total_rows + 1,  # +1 for header
                       unit='rows')

            for row in pbar:
                csv_writer.writerow(row)

            pbar.close()

        wb.close()
        logger.info(f"Conversion complete: {temp_csv}")

    def load(self) -> pd.DataFrame:
        """
        Load Excel file with chunked reading and real progress tracking.

        Strategy for large files (>50MB):
        1. Convert XLSX → CSV (faster I/O, real progress)
        2. Load CSV in chunks with tqdm
        3. Concatenate chunks into final DataFrame

        Returns:
            Loaded DataFrame

        Raises:
            FileNotFoundError: If file doesn't exist
            ValueError: If file cannot be read
        """
        if not self.file_path.exists():
            raise FileNotFoundError(f"File not found: {self.file_path}")

        file_size_mb = self.file_path.stat().st_size / (1024 * 1024)
        logger.info(f"Loading Excel file: {self.file_path} ({file_size_mb:.1f} MB)")

        try:
            # For large files (>50MB), use optimized chunked reading
            if file_size_mb > 50:
                logger.info("Large file detected - using optimized chunked loading")
                self.df = self._load_large_file()
            else:
                # For small files, use standard pandas read_excel
                logger.info("Small file - using standard loading")
                with tqdm(total=1, desc="Loading Excel") as pbar:
                    self.df = pd.read_excel(self.file_path, engine='openpyxl')
                    pbar.update(1)

            logger.info(f"Loaded {len(self.df)} rows, {len(self.df.columns)} columns")

            # Cleanup temporary CSV if exists
            self._cleanup_temp_csv()

            return self.df

        except Exception as e:
            self._cleanup_temp_csv()
            raise ValueError(f"Failed to load Excel: {str(e)}")

    def _load_large_file(self) -> pd.DataFrame:
        """
        Load large Excel file using CSV conversion and chunked reading.

        Returns:
            Loaded DataFrame with all rows
        """
        # Step 1: Get total rows for progress tracking
        total_rows = self._get_total_rows()
        logger.info(f"Estimated rows: {total_rows:,}")

        # Step 2: Convert XLSX to temporary CSV
        temp_dir = Path(tempfile.gettempdir())
        self._temp_csv_path = temp_dir / f"excel_loader_{self.file_path.stem}.csv"

        self._convert_xlsx_to_csv(self._temp_csv_path, total_rows)

        # Step 3: Load CSV in chunks with progress
        logger.info(f"Loading CSV in chunks of {self.chunk_size:,} rows...")

        chunks = []
        chunk_iterator = pd.read_csv(
            self._temp_csv_path,
            chunksize=self.chunk_size,
            encoding='utf-8',
            low_memory=False  # Prevent dtype warnings
        )

        # Progress bar for reading chunks
        total_chunks = (total_rows // self.chunk_size) + 1 if total_rows > 0 else None
        pbar = tqdm(chunk_iterator,
                   desc="Reading chunks",
                   total=total_chunks,
                   unit='chunk')

        for chunk in pbar:
            chunks.append(chunk)
            pbar.set_postfix({'rows': len(chunk)})

        pbar.close()

        # Step 4: Concatenate all chunks
        logger.info(f"Concatenating {len(chunks)} chunks...")
        df = pd.concat(chunks, ignore_index=True)

        logger.info(f"Successfully loaded {len(df):,} rows")
        return df

    def _cleanup_temp_csv(self) -> None:
        """Remove temporary CSV file if it exists."""
        if self._temp_csv_path and self._temp_csv_path.exists():
            try:
                self._temp_csv_path.unlink()
                logger.debug(f"Cleaned up temporary CSV: {self._temp_csv_path}")
            except Exception as e:
                logger.warning(f"Could not remove temporary CSV: {e}")

    def validate(self) -> bool:
        """
        Validate loaded data.

        Returns:
            True if valid

        Raises:
            ValueError: If validation fails
        """
        if self.df is None:
            raise ValueError("No data loaded. Call load() first.")

        # Check required columns
        missing = [col for col in self.REQUIRED_COLUMNS if col not in self.df.columns]
        if missing:
            raise ValueError(f"Missing required columns: {missing}")

        # Check data types
        if not pd.api.types.is_numeric_dtype(self.df['Ресурс | Стоимость (руб.)']):
            logger.warning("Converting 'Ресурс | Стоимость (руб.)' to numeric")
            self.df['Ресурс | Стоимость (руб.)'] = pd.to_numeric(
                self.df['Ресурс | Стоимость (руб.)'], errors='coerce'
            )

        # Check critical NaN values
        for col in self.CRITICAL_COLUMNS:
            nan_count = self.df[col].isna().sum()
            if nan_count > 0:
                raise ValueError(f"Column '{col}' has {nan_count} NaN values")

        logger.info("Validation passed")
        return True

    def get_statistics(self) -> Dict[str, Any]:
        """
        Get statistics about loaded data.

        Returns:
            Dict with total_rows, unique_rates, row_types
        """
        if self.df is None:
            raise ValueError("No data loaded. Call load() first.")

        stats = {
            'total_rows': len(self.df),
            'unique_rates': self.df['Расценка | Код'].nunique(),
            'row_types': self.df['Тип строки'].value_counts().to_dict()
        }

        logger.info(f"Statistics: {stats}")
        return stats
