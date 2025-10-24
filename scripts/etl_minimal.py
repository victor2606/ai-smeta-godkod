#!/usr/bin/env python3
"""
Minimal memory-efficient ETL for large Excel files.
Reads XLSX row-by-row and writes directly to SQLite in small batches.
"""

import sys
import sqlite3
import logging
from pathlib import Path
from openpyxl import load_workbook
from tqdm import tqdm
from collections import defaultdict

logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)

# Column mapping (0-indexed) - correct mapping based on actual Excel structure
COLUMN_MAPPING = {
    "rate_code": 13,  # Расценка | Код
    "rate_full_name": 14,  # Расценка | Исходное наименование
    "unit_type": 16,  # Расценка | Ед. изм.
    "row_type": 17,  # Тип строки
    "resource_code": 19,  # Ресурс | Код
    "resource_cost": 26,  # Ресурс | Стоимость (руб.)
    "median_price": 27,  # Прайс | АбстРесурс | Сметная цена текущая_median
}


def create_database(db_path: Path):
    """Create SQLite database with schema."""
    logger.info(f"Creating database: {db_path}")

    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    # Create tables
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS rates (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            rate_code TEXT NOT NULL UNIQUE,
            rate_full_name TEXT NOT NULL,
            unit_type TEXT NOT NULL,
            total_cost REAL DEFAULT 0.0,
            labor_cost REAL DEFAULT 0.0,
            machine_cost REAL DEFAULT 0.0,
            material_cost REAL DEFAULT 0.0,
            unit_quantity REAL DEFAULT 1.0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS resources (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            rate_code TEXT NOT NULL,
            resource_code TEXT NOT NULL,
            resource_cost REAL NOT NULL,
            median_price REAL,
            FOREIGN KEY (rate_code) REFERENCES rates(rate_code) ON DELETE CASCADE
        )
    """)

    # Create FTS5 table
    cursor.execute("""
        CREATE VIRTUAL TABLE IF NOT EXISTS rates_fts USING fts5(
            rate_code,
            rate_full_name,
            content=rates,
            content_rowid=id
        )
    """)

    # Create indexes
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_rates_code ON rates(rate_code)")
    cursor.execute(
        "CREATE INDEX IF NOT EXISTS idx_resources_rate ON resources(rate_code)"
    )
    cursor.execute(
        "CREATE INDEX IF NOT EXISTS idx_resources_code ON resources(resource_code)"
    )

    conn.commit()
    return conn


def process_excel_streaming(
    excel_path: Path, db_conn: sqlite3.Connection, batch_size: int = 1000
):
    """
    Process Excel file row-by-row with minimal memory usage.

    Strategy:
    1. Read Excel in read_only mode (no worksheet caching)
    2. Process rows in small batches
    3. Aggregate rate data on-the-fly
    4. Insert to SQLite in batches
    """
    logger.info(f"Opening Excel file: {excel_path}")

    # Load workbook in read-only mode (streaming)
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active

    # Get total rows for progress bar
    total_rows = ws.max_row
    logger.info(f"Total rows: {total_rows:,}")

    # Data structures
    rates_dict = {}  # rate_code -> {costs, name, unit}
    resources_batch = []

    # Read header (first row) - skip it
    header_iter = ws.iter_rows(min_row=1, max_row=1, values_only=True)
    header = next(header_iter)
    logger.info(f"Columns: {len(header)}")

    # Process data rows
    cursor = db_conn.cursor()
    processed_rows = 0

    pbar = tqdm(
        ws.iter_rows(min_row=2, values_only=True),
        desc="Processing rows",
        total=total_rows - 1,
        unit="rows",
    )

    for row in pbar:
        try:
            # Extract values
            rate_code = row[COLUMN_MAPPING["rate_code"]]
            rate_name = row[COLUMN_MAPPING["rate_full_name"]]
            unit_type = row[COLUMN_MAPPING["unit_type"]]
            row_type = row[COLUMN_MAPPING["row_type"]]
            resource_code = row[COLUMN_MAPPING["resource_code"]]
            resource_cost = row[COLUMN_MAPPING["resource_cost"]]
            median_price = row[COLUMN_MAPPING["median_price"]]

            # Skip rows without rate code
            if not rate_code:
                continue

            # Initialize rate if first time seeing it
            if rate_code not in rates_dict:
                rates_dict[rate_code] = {
                    "name": rate_name or "",
                    "unit": unit_type or "",
                    "total_cost": 0.0,
                    "labor_cost": 0.0,
                    "machine_cost": 0.0,
                    "material_cost": 0.0,
                }

            # Aggregate costs based on row type
            if row_type and resource_cost is not None:
                cost = float(resource_cost) if resource_cost else 0.0
                rates_dict[rate_code]["total_cost"] += cost

                if "труд" in str(row_type).lower() or "рабоч" in str(row_type).lower():
                    rates_dict[rate_code]["labor_cost"] += cost
                elif (
                    "машин" in str(row_type).lower() or "механ" in str(row_type).lower()
                ):
                    rates_dict[rate_code]["machine_cost"] += cost
                elif "материал" in str(row_type).lower():
                    rates_dict[rate_code]["material_cost"] += cost

            # Add resource record
            if resource_code:
                resources_batch.append(
                    (
                        rate_code,
                        resource_code,
                        float(resource_cost) if resource_cost else 0.0,
                        float(median_price) if median_price else None,
                    )
                )

            processed_rows += 1

            # Batch insert resources
            if len(resources_batch) >= batch_size:
                cursor.executemany(
                    "INSERT INTO resources (rate_code, resource_code, resource_cost, median_price) VALUES (?, ?, ?, ?)",
                    resources_batch,
                )
                db_conn.commit()
                resources_batch = []
                pbar.set_postfix(
                    {"rates": len(rates_dict), "resources": processed_rows}
                )

        except Exception as e:
            logger.warning(f"Error processing row: {e}")
            continue

    pbar.close()
    wb.close()

    # Insert remaining resources
    if resources_batch:
        cursor.executemany(
            "INSERT INTO resources (rate_code, resource_code, resource_cost, median_price) VALUES (?, ?, ?, ?)",
            resources_batch,
        )
        db_conn.commit()

    # Insert aggregated rates
    logger.info(f"Inserting {len(rates_dict):,} aggregated rates...")
    rates_batch = []
    for rate_code, data in tqdm(rates_dict.items(), desc="Inserting rates"):
        rates_batch.append(
            (
                rate_code,
                data["name"],
                data["unit"],
                data["total_cost"],
                data["labor_cost"],
                data["machine_cost"],
                data["material_cost"],
                1.0,  # unit_quantity
            )
        )

        if len(rates_batch) >= batch_size:
            cursor.executemany(
                """INSERT INTO rates
                   (rate_code, rate_full_name, unit_type, total_cost, labor_cost, machine_cost, material_cost, unit_quantity)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                rates_batch,
            )
            db_conn.commit()
            rates_batch = []

    if rates_batch:
        cursor.executemany(
            """INSERT INTO rates
               (rate_code, rate_full_name, unit_type, total_cost, labor_cost, machine_cost, material_cost, unit_quantity)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
            rates_batch,
        )
        db_conn.commit()

    # Populate FTS5 index
    logger.info("Building FTS5 search index...")
    cursor.execute("""
        INSERT INTO rates_fts (rowid, rate_code, rate_full_name)
        SELECT id, rate_code, rate_full_name FROM rates
    """)
    db_conn.commit()

    logger.info(f"✅ Processed {processed_rows:,} rows")
    logger.info(f"✅ Created {len(rates_dict):,} rates")


def verify_database(db_path: Path):
    """Verify database contents."""
    logger.info("Verifying database...")
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    rates_count = cursor.execute("SELECT COUNT(*) FROM rates").fetchone()[0]
    resources_count = cursor.execute("SELECT COUNT(*) FROM resources").fetchone()[0]
    fts_count = cursor.execute("SELECT COUNT(*) FROM rates_fts").fetchone()[0]

    logger.info(f"✅ Rates: {rates_count:,}")
    logger.info(f"✅ Resources: {resources_count:,}")
    logger.info(f"✅ FTS index: {fts_count:,} entries")

    db_size_mb = db_path.stat().st_size / (1024 * 1024)
    logger.info(f"✅ Database size: {db_size_mb:.1f} MB")

    conn.close()


def main():
    if len(sys.argv) < 3:
        print("Usage: python etl_minimal.py <excel_file> <output_db>")
        sys.exit(1)

    excel_path = Path(sys.argv[1])
    db_path = Path(sys.argv[2])

    if not excel_path.exists():
        logger.error(f"Excel file not found: {excel_path}")
        sys.exit(1)

    # Ensure output directory exists
    db_path.parent.mkdir(parents=True, exist_ok=True)

    # Remove old database if exists
    if db_path.exists():
        logger.info(f"Removing old database: {db_path}")
        db_path.unlink()

    logger.info("=" * 60)
    logger.info("Starting Minimal Memory-Efficient ETL")
    logger.info("=" * 60)

    # Create database
    conn = create_database(db_path)

    # Process Excel
    process_excel_streaming(excel_path, conn, batch_size=1000)

    # Verify
    conn.close()
    verify_database(db_path)

    logger.info("=" * 60)
    logger.info("🎉 ETL Completed Successfully!")
    logger.info("=" * 60)


if __name__ == "__main__":
    main()
